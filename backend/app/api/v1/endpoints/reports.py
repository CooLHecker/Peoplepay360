"""Admin-only downloadable reports.

Per idea.md's Employee -> Contract -> ... -> Payrun chain and
corrections.md section 43 ("Add a Reports section if the existing
backend/data supports it"), this starts with a single report that the
current data model can actually back: an Employee Summary combining
real Employee, Contract, and TimeOffRequest rows.

Attendance has no real data model yet (see app/api/v1/endpoints/
attendance.py - it's a hardcoded stub), so the attendance columns here
are clearly-labelled sample data, not a claim about real check-ins.
Once a real Attendance model exists, ``_sample_attendance`` is the only
function that needs to be replaced with a real query.
"""

import io
import random
from calendar import monthrange
from datetime import date, datetime

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func
from sqlalchemy.orm import Session, selectinload

from app.api.deps import require_roles
from app.db.session import get_db as db_dependency
from app.models import Contract, Employee, RoleName, TimeOffRequest, TimeOffStatus

router = APIRouter()

# System-wide reports are Admin territory (idea.md section 9: "Admin -
# Complete system access"), not shared with the HR/payroll roles that
# read individual modules.
_REPORT_ROLES = (RoleName.ADMIN,)

_BRAND_HEX = "714B67"


def _active_contract(contracts: list[Contract], on: date) -> Contract | None:
    """Pick the contract that applies "on" a given date.

    Mirrors the resolution rule in endpoints/contracts.py
    (_effective_status): not-cancelled, started, and not yet ended.
    Falls back to the most recently started contract if none are
    currently running, so a terminated employee's last role still
    shows up in the report instead of a blank row.
    """
    running = [
        c
        for c in contracts
        if c.status.value != "cancelled" and c.start_date <= on and (c.end_date is None or c.end_date >= on)
    ]
    if running:
        return max(running, key=lambda c: c.start_date)
    if contracts:
        return max(contracts, key=lambda c: c.start_date)
    return None


def _sample_attendance(employee_id: int, working_days: int) -> tuple[int, int, int]:
    """Deterministic placeholder attendance for one employee.

    Seeded on employee_id so the same employee always gets the same
    numbers within a run (and across xlsx/pdf for the same request),
    rather than random noise on every call. Replace this function with
    a real query once an Attendance model/table exists.
    """
    if working_days <= 0:
        return 0, 0, 0
    rng = random.Random(employee_id * 7919)
    present = round(working_days * rng.uniform(0.85, 1.0))
    present = min(present, working_days)
    late = rng.randint(0, min(3, working_days - present)) if working_days - present > 0 else 0
    absent = working_days - present - late
    return present, late, absent


def _working_days_so_far(today: date) -> int:
    """Weekday count from day 1 of the current month through today."""
    days_in_month = monthrange(today.year, today.month)[1]
    return sum(
        1
        for day in range(1, min(today.day, days_in_month) + 1)
        if date(today.year, today.month, day).weekday() < 5
    )


def _build_rows(db: Session) -> list[dict]:
    today = date.today()
    working_days = _working_days_so_far(today)
    year_start = date(today.year, 1, 1)

    employees = (
        db.query(Employee)
        .options(selectinload(Employee.contracts))
        .order_by(Employee.id.asc())
        .all()
    )

    # One aggregate query for approved time-off days this year, and one
    # for pending requests, instead of N+1 queries per employee.
    approved_days_by_employee = dict(
        db.query(TimeOffRequest.employee_id, func.coalesce(func.sum(TimeOffRequest.number_of_days), 0))
        .filter(TimeOffRequest.status == TimeOffStatus.approved, TimeOffRequest.start_date >= year_start)
        .group_by(TimeOffRequest.employee_id)
        .all()
    )
    pending_count_by_employee = dict(
        db.query(TimeOffRequest.employee_id, func.count(TimeOffRequest.id))
        .filter(TimeOffRequest.status == TimeOffStatus.submitted)
        .group_by(TimeOffRequest.employee_id)
        .all()
    )

    rows = []
    for employee in employees:
        contract = _active_contract(list(employee.contracts), today)
        present, late, absent = _sample_attendance(employee.id, working_days)
        rows.append(
            {
                "id": employee.id,
                "full_name": employee.full_name,
                "work_email": employee.work_email or "—",
                "employment_status": employee.employment_status.value,
                "department": contract.department if contract else "—",
                "job_position": contract.job_position if contract else "—",
                "wage": float(contract.wage) if contract else None,
                "contract_status": contract.status.value if contract else "no contract",
                "approved_time_off_days": float(approved_days_by_employee.get(employee.id, 0)),
                "pending_time_off_requests": int(pending_count_by_employee.get(employee.id, 0)),
                "attendance_present": present,
                "attendance_late": late,
                "attendance_absent": absent,
            }
        )
    return rows


_ATTENDANCE_NOTE = (
    "Attendance columns are sample placeholder data (no attendance check-in/out "
    "records exist yet) and do not reflect real check-ins. All other columns "
    "reflect live employee, contract, and time-off data."
)


def _build_xlsx(rows: list[dict], generated_at: datetime) -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Employee Summary"

    headers = [
        "Employee ID",
        "Full Name",
        "Work Email",
        "Employment Status",
        "Department",
        "Job Position",
        "Wage (INR)",
        "Contract Status",
        "Approved Time Off (days, YTD)",
        "Pending Time Off Requests",
        "Attendance Present (sample)",
        "Attendance Late (sample)",
        "Attendance Absent (sample)",
    ]
    ws.append(headers)
    header_fill = PatternFill(start_color=_BRAND_HEX, end_color=_BRAND_HEX, fill_type="solid")
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left")

    for row in rows:
        ws.append(
            [
                row["id"],
                row["full_name"],
                row["work_email"],
                row["employment_status"],
                row["department"],
                row["job_position"],
                row["wage"],
                row["contract_status"],
                row["approved_time_off_days"],
                row["pending_time_off_requests"],
                row["attendance_present"],
                row["attendance_late"],
                row["attendance_absent"],
            ]
        )

    for col_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(header) * 1.1)

    ws.freeze_panes = "A2"

    note_row = ws.max_row + 2
    ws.cell(row=note_row, column=1, value=f"Generated {generated_at.strftime('%Y-%m-%d %H:%M')}").font = Font(italic=True, size=9)
    note_cell = ws.cell(row=note_row + 1, column=1, value=_ATTENDANCE_NOTE)
    note_cell.font = Font(italic=True, size=9, color="A36B12")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _build_pdf(rows: list[dict], generated_at: datetime) -> io.BytesIO:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import landscape, letter
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import inch
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError:
        return _build_pdf_without_dependencies(rows, generated_at)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=0.4 * inch,
        rightMargin=0.4 * inch,
        topMargin=0.5 * inch,
        bottomMargin=0.5 * inch,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("ReportTitle", parent=styles["Title"], textColor=colors.HexColor(f"#{_BRAND_HEX}"))
    note_style = ParagraphStyle("Note", parent=styles["Normal"], fontSize=8, textColor=colors.HexColor("#A36B12"))
    meta_style = ParagraphStyle("Meta", parent=styles["Normal"], fontSize=8, textColor=colors.grey)

    header = [
        "ID",
        "Full Name",
        "Work Email",
        "Status",
        "Department",
        "Position",
        "Wage",
        "Contract",
        "Approved TO (YTD)",
        "Pending TO",
        "Present*",
        "Late*",
        "Absent*",
    ]
    data = [header]
    for row in rows:
        data.append(
            [
                str(row["id"]),
                row["full_name"],
                row["work_email"],
                row["employment_status"],
                row["department"],
                row["job_position"],
                f"{row['wage']:,.0f}" if row["wage"] is not None else "—",
                row["contract_status"],
                f"{row['approved_time_off_days']:g}",
                str(row["pending_time_off_requests"]),
                str(row["attendance_present"]),
                str(row["attendance_late"]),
                str(row["attendance_absent"]),
            ]
        )

    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{_BRAND_HEX}")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 7.5),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F2F6")]),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E6E0E5")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )

    story = [
        Paragraph("Employee Summary Report", title_style),
        Paragraph(f"Generated {generated_at.strftime('%Y-%m-%d %H:%M')} · {len(rows)} employees", meta_style),
        Spacer(1, 12),
        table,
        Spacer(1, 10),
        Paragraph(f"* {_ATTENDANCE_NOTE}", note_style),
    ]
    doc.build(story)
    buffer.seek(0)
    return buffer


def _pdf_text(value: object) -> str:
    return str(value).encode("ascii", "replace").decode("ascii")


def _build_pdf_without_dependencies(rows: list[dict], generated_at: datetime) -> io.BytesIO:
    """Create a valid, readable PDF when optional report libraries are absent."""
    lines = [
        "Employee Summary Report",
        f"Generated {generated_at.strftime('%Y-%m-%d %H:%M')} - {len(rows)} employees",
        "",
        "ID | Full Name | Work Email | Status | Department | Position | Wage | Contract",
    ]
    for row in rows:
        wage = f"{row['wage']:,.0f}" if row["wage"] is not None else "-"
        lines.append(
            " | ".join(
                _pdf_text(value)
                for value in (
                    row["id"],
                    row["full_name"],
                    row["work_email"],
                    row["employment_status"],
                    row["department"],
                    row["job_position"],
                    wage,
                    row["contract_status"],
                )
            )
        )
    lines.extend(["", _pdf_text(_ATTENDANCE_NOTE)])

    pages = [lines[index:index + 42] for index in range(0, len(lines), 42)] or [[]]
    objects: list[str] = []
    page_ids: list[int] = []
    content_ids: list[int] = []

    objects.append("<< /Type /Catalog /Pages 2 0 R >>")
    objects.append("")
    for page_lines in pages:
        content = ["BT", "/F1 8 Tf", "40 560 Td"]
        for index, line in enumerate(page_lines):
            if index:
                content.append("0 -13 Td")
            escaped = _pdf_text(line).replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")
            content.append(f"({escaped[:180]}) Tj")
        content.append("ET")
        content_value = "\n".join(content)
        content_ids.append(len(objects) + 1)
        objects.append(f"<< /Length {len(content_value.encode('ascii'))} >>\nstream\n{content_value}\nendstream")
        page_ids.append(len(objects) + 1)
        objects.append("")

    objects[1] = f"<< /Type /Pages /Kids [{' '.join(f'{page_id} 0 R' for page_id in page_ids)}] /Count {len(page_ids)} >>"
    for index, page_id in enumerate(page_ids):
        objects[page_id - 1] = (
            f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 612] "
            f"/Resources << /Font << /F1 {len(objects) + 1} 0 R >> >> "
            f"/Contents {content_ids[index]} 0 R >>"
        )
    objects.append("<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>")

    pdf = bytearray(b"%PDF-1.4\n")
    offsets = [0]
    for object_id, value in enumerate(objects, start=1):
        offsets.append(len(pdf))
        pdf.extend(f"{object_id} 0 obj\n{value}\nendobj\n".encode("ascii"))
    xref_offset = len(pdf)
    pdf.extend(f"xref\n0 {len(objects) + 1}\n0000000000 65535 f \n".encode("ascii"))
    pdf.extend("".join(f"{offset:010d} 00000 n \n" for offset in offsets[1:]).encode("ascii"))
    pdf.extend(
        f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref_offset}\n%%EOF".encode("ascii")
    )
    return io.BytesIO(pdf)


@router.get("/employees-summary", dependencies=[Depends(require_roles(*_REPORT_ROLES))])
def employees_summary_report(
    format: str = Query(default="xlsx", pattern="^(xlsx|pdf)$"),
    db: Session = Depends(db_dependency),
) -> StreamingResponse:
    rows = _build_rows(db)
    generated_at = datetime.now()
    stamp = generated_at.strftime("%Y%m%d-%H%M")

    if format == "pdf":
        buffer = _build_pdf(rows, generated_at)
        media_type = "application/pdf"
        filename = f"employee-summary-report-{stamp}.pdf"
    else:
        buffer = _build_xlsx(rows, generated_at)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = f"employee-summary-report-{stamp}.xlsx"

    return StreamingResponse(
        buffer,
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
